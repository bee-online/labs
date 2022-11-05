"""Python program to demonstrate
selenium"""


import openpyxl
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
import time

driver = webdriver.Firefox()
actions = ActionChains(driver)
driver.maximize_window()
driver.implicitly_wait(0.5)

#Main Function starts here
#
driver.get("http://iradmission.bisefsd.edu.pk/Account/Login.aspx")

# # controls on website
username = driver.find_element_by_id('ContentPlaceHolder1_LoginUser_UserName')
username.send_keys("#Write Institute username here")
password = driver.find_element_by_id('ContentPlaceHolder1_LoginUser_Password')
password.send_keys("Write Institute password here")
driver.find_element_by_id('ContentPlaceHolder1_LoginUser_Button1').click()

driver.implicitly_wait(1.5)
driver.find_element_by_xpath('/html/body/form/div[3]/div[1]/div/div/div[1]/button').click()
print("Xpath close button")
driver.implicitly_wait(1.5)

#popup
# close_popup = driver.find_element_by_css_selector(".close")
# driver.execute_script("arguments[0].scrollIntoView();", close_popup)
#Close popup
# driver.find_element_by_css_selector(".close").click()

#Start work
driver.get("http://iradmission.bisefsd.edu.pk/P2AdmissionForms/AdmissionListP2.aspx")
driver.get("http://iradmission.bisefsd.edu.pk/P2AdmissionForms/AddNew.aspx")
panel_body = driver.find_element_by_id("ContentPlaceHolder1_txtLastRollNo")

#import roll no column from excel file
path = "C:\\Users\\MAHBOOBALAMMT-SCI-dF\\Desktop\\roll.xlsx"
if path:
    wb_obj = openpyxl.load_workbook(path)
    print("Excel file loaded!")

sheet_obj = wb_obj.active
max_row = sheet_obj.max_row

rolls = []
grades= []
# Loop will print all rows name - ROLL Nos
for i in range(1, max_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    if "ROLL" in str(cell_obj.value):
        continue
    else:
        rolls.append(str(cell_obj.value))
print("Printing List ", rolls)

# grades section
for x in range(1, max_row + 1):
    cell_obj = sheet_obj.cell(row = x, column = 2)
    string_grade = str(cell_obj.value)
    if string_grade == "GRADE":
        continue
    else:
        grades.append(string_grade)
print("Printing Grades ", grades)
# print("Printing 1st element:", grades[0])

#Getting each value from list
for count, val in enumerate(rolls):
    panel_body = driver.find_element_by_id("ContentPlaceHolder1_txtLastRollNo")
    print("Printing count variable", count)
    driver.implicitly_wait(1)

    panel_body.send_keys(val)
    print("Got" +val+"Value")
    driver.find_element_by_id('ContentPlaceHolder1_btnGet').click()
    driver.implicitly_wait(1)
    dropdown = driver.find_element_by_id("ContentPlaceHolder1_ddlDistrict")
    driver.execute_script("arguments[0].scrollIntoView();", dropdown)

    driver.implicitly_wait(0.5)
    driver.find_element_by_xpath(
        "//select[@id='ContentPlaceHolder1_ddlDistrict']/option[text()='TTSINGH']").click()
    driver.implicitly_wait(1)
    driver.find_element_by_xpath(
        "//select[@id='ContentPlaceHolder1_ddlTehsil']/option[text()='PIR MAHAL']").click()

    driver.implicitly_wait(1)

    driver.find_element_by_xpath("//select[@id='ContentPlaceHolder1_ddlIGrade']/option[@value='"+grades[count]+"']").click()
    driver.find_element_by_id('ContentPlaceHolder1_ButtonSave').click()

    driver.implicitly_wait(1)
    driver.get("http://iradmission.bisefsd.edu.pk/P2AdmissionForms/AddNew.aspx")

    print("Save Clicked")
    # break



# panel_body.send_keys(cell_obj.value)
# driver.find_element_by_id('ContentPlaceHolder1_btnGet').click()




